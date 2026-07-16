"""Memory-efficient profiler for the PAPL transposed Excel workbook."""

from __future__ import annotations

import argparse
import html
import json
import math
import re
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots


DEFAULT_INPUT = Path("data/PAPL Phase 1- Transposed File_3_vs.xlsx")
DEFAULT_OUTPUT = Path("docs/data_profile.json")
DEFAULT_SHEET = "Final Ver"
TOP_N = 15
CHART_DIR = Path("docs/interactive_charts")

STATE_NAMES = {
    "AP": "Andhra Pradesh", "DEL": "Delhi", "GUJ": "Gujarat",
    "HAR": "Haryana", "KAR": "Karnataka", "MAH": "Maharashtra",
    "PUNJ": "Punjab", "PUJ": "Punjab", "RAJ": "Rajasthan",
    "TLG": "Telangana", "TN": "Tamil Nadu", "UP": "Uttar Pradesh",
    "WB": "West Bengal",
}
METRICS = [
    "Sales Value (000)", "Sales (VOL UNIT CASES)", "Sales Units",
    "Share of Sales Value - Product", "Numeric Distribution Handling (C)",
    "Numeric Out of Stock (C)", "Wghtd Dist Out of Stock",
    "Number of Stores Retailing (C)", "Wghtd Dist Handling (C) - CATEGORY",
    "Price per Sales (VOL UNIT CASES)", "Price per Sales Unit",
    "Stocks (VOL UNIT CASES)", "Purchase (VOL UNIT CASES)",
]


def clean(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def profile_sheet(ws) -> dict[str, Any]:
    rows = ws.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else f"Unnamed_{i + 1}" for i, value in enumerate(next(rows))]
    columns = {
        header: {
            "non_null": 0,
            "numeric": 0,
            "text": 0,
            "errors": 0,
            "numeric_sum": 0.0,
            "numeric_min": None,
            "numeric_max": None,
        }
        for header in headers
    }
    tracked = {
        header: Counter()
        for header in headers
        if header in {
            "Markets", "S4CATEGORY", "MANUFACTURER", "ITEM", "Pack Type",
            "Pack Type_Size", "Size", "Addition", "Attribute", "ITEM - Pack Size"
        }
    }
    row_count = 0
    empty_rows = 0
    populated_cells = Counter()
    error_values = Counter()

    for row_count, row in enumerate(rows, start=1):
        populated = 0
        for index, header in enumerate(headers):
            value = row[index] if index < len(row) else None
            if value is None or value == "":
                continue
            populated += 1
            stat = columns[header]
            stat["non_null"] += 1
            if isinstance(value, bool):
                stat["text"] += 1
            elif isinstance(value, (int, float)) and math.isfinite(float(value)):
                number = float(value)
                stat["numeric"] += 1
                stat["numeric_sum"] += number
                stat["numeric_min"] = number if stat["numeric_min"] is None else min(stat["numeric_min"], number)
                stat["numeric_max"] = number if stat["numeric_max"] is None else max(stat["numeric_max"], number)
            else:
                text = str(clean(value)).strip()
                stat["text"] += 1
                if text.startswith("#"):
                    stat["errors"] += 1
                    error_values[(header, text)] += 1
            if header in tracked:
                tracked[header][str(clean(value))] += 1
        populated_cells[populated] += 1
        if populated == 0:
            empty_rows += 1

    for stat in columns.values():
        stat["missing"] = row_count - stat["non_null"]
        stat["missing_pct"] = round(100 * stat["missing"] / row_count, 2) if row_count else 0
        stat["numeric_sum"] = round(stat["numeric_sum"], 4)

    return {
        "rows": row_count,
        "columns": len(headers),
        "empty_rows": empty_rows,
        "column_profile": columns,
        "top_values": {
            header: [{"value": value, "count": count} for value, count in counter.most_common(TOP_N)]
            for header, counter in tracked.items()
        },
        "unique_counts": {header: len(counter) for header, counter in tracked.items()},
        "error_values": [
            {"column": column, "value": value, "count": count}
            for (column, value), count in error_values.most_common()
        ],
        "populated_cells_per_row": dict(sorted(populated_cells.items())),
    }


def load_final_data(path: Path, sheet: str) -> pd.DataFrame:
    """Load only populated rows, avoiding the sheet's 884k formatted blank rows."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet]
    iterator = worksheet.iter_rows(values_only=True)
    headers = list(next(iterator))
    records = [row for row in iterator if any(value is not None and value != "" for value in row)]
    workbook.close()
    frame = pd.DataFrame.from_records(records, columns=headers)
    for column in METRICS + ["Size"]:
        if column in frame:
            frame[column] = pd.to_numeric(frame[column], errors="coerce")
    frame["State"] = frame["Markets"].str.extract(r"/([^/]+)/INDIA$")[0].map(STATE_NAMES)
    frame["Brand"] = (
        frame["ITEM"].fillna("").str.lstrip("#").str.split("\\", regex=False).str[0].str.strip()
    )
    frame.loc[frame["Brand"].eq(""), "Brand"] = pd.NA
    frame["Hierarchy"] = "Category total"
    frame.loc[frame["MANUFACTURER"].notna(), "Hierarchy"] = "Parle aggregate"
    frame.loc[frame["ITEM"].notna(), "Hierarchy"] = "SKU"
    # Pack Type is corrected in Final Ver from the authoritative Pack Type_Size prefix.
    frame["Derived Pack Type"] = frame["Pack Type"].fillna("Unknown")
    is_tetra = frame["Derived Pack Type"].eq("TPK")
    frame["Pack Bucket"] = pd.cut(
        frame["Size"], bins=[-math.inf, 200, 300, 600, math.inf],
        labels=["Individual / On-the-go (≤200ml)", "Single-serve (201–300ml)",
                "Sharing (301–600ml)", "Family / Party (>600ml)"],
    ).astype("object")
    frame.loc[is_tetra, "Pack Bucket"] = "Tetra Pack"
    frame.loc[is_tetra & frame["Size"].eq(80), "Pack Bucket"] = "Tetra Pack – 80ml"
    frame["Pack Bucket"] = frame["Pack Bucket"].fillna("Unknown")
    return frame


def sum_metric(frame: pd.DataFrame, groups: list[str], metric: str) -> pd.DataFrame:
    subset = frame.dropna(subset=groups + [metric])
    return subset.groupby(groups, as_index=False, observed=True)[metric].sum()


def max_metric(frame: pd.DataFrame, groups: list[str], metric: str) -> pd.DataFrame:
    subset = frame.dropna(subset=groups + [metric])
    return subset.groupby(groups, as_index=False, observed=True)[metric].max()


def save_chart(fig: go.Figure, number: int, slug: str, title: str, chart_dir: Path) -> dict[str, str]:
    filename = f"{number:02d}_{slug}.html"
    image_filename = f"{number:02d}_{slug}.png"
    fig.update_layout(
        template="plotly_white", title={"text": title, "x": 0.02},
        font={"family": "Arial", "size": 13}, hoverlabel={"namelength": -1},
        margin={"l": 70, "r": 40, "t": 90, "b": 70},
    )
    fig.add_annotation(
        text="Source: Final Ver, Apr–Sep 2025. Values retain source units.",
        xref="paper", yref="paper", x=0, y=-0.18, showarrow=False,
        font={"size": 10, "color": "#666"}, align="left",
    )
    fig.write_html(chart_dir / filename, include_plotlyjs="directory", full_html=True)
    fig.write_image(chart_dir / image_filename, width=1400, height=800, scale=1)
    return {"file": filename, "image": image_filename, "title": title}


def create_interactive_charts(frame: pd.DataFrame, chart_dir: Path) -> list[dict[str, str]]:
    chart_dir.mkdir(parents=True, exist_ok=True)
    category = frame[frame["Hierarchy"].eq("Category total")]
    parle = frame[frame["Hierarchy"].eq("Parle aggregate")]
    sku = frame[frame["Hierarchy"].eq("SKU")]
    sales = "Sales Value (000)"
    nd = "Numeric Distribution Handling (C)"
    stores = "Number of Stores Retailing (C)"
    charts: list[dict[str, str]] = []

    # 1. National category opportunity and Parle share.
    cat_value = sum_metric(category, ["S4CATEGORY"], sales).rename(columns={sales: "Category Value"})
    pa_value = sum_metric(parle, ["S4CATEGORY"], sales).rename(columns={sales: "Parle Value"})
    national = cat_value.merge(pa_value, on="S4CATEGORY", how="left").fillna({"Parle Value": 0})
    national["Parle Share %"] = 100 * national["Parle Value"] / national["Category Value"]
    national = national.sort_values("Category Value", ascending=False)
    fig = make_subplots(specs=[[{"secondary_y": True}]])
    fig.add_bar(x=national["S4CATEGORY"], y=national["Category Value"], name="Category value", marker_color="#17365D", secondary_y=False)
    fig.add_scatter(x=national["S4CATEGORY"], y=national["Parle Share %"], name="Parle value share %", mode="lines+markers+text", text=national["Parle Share %"].round(1), textposition="top center", marker_color="#F5A623", secondary_y=True)
    fig.update_yaxes(title_text="Category value (source units)", secondary_y=False)
    fig.update_yaxes(title_text="Parle share (%)", secondary_y=True)
    charts.append(save_chart(fig, 1, "national_category_opportunity", "National Category Opportunity and Parle Share", chart_dir))

    # 2. State opportunity and current share.
    state_cat = sum_metric(category, ["State"], sales).rename(columns={sales: "Category Value"})
    state_pa = sum_metric(parle, ["State"], sales).rename(columns={sales: "Parle Value"})
    state = state_cat.merge(state_pa, on="State", how="left").fillna({"Parle Value": 0})
    state["Parle Share %"] = 100 * state["Parle Value"] / state["Category Value"]
    state = state.sort_values("Category Value", ascending=True)
    fig = px.bar(state, x="Category Value", y="State", orientation="h", color="Parle Share %", color_continuous_scale="Blues", hover_data={"Parle Value": ":,.2f", "Parle Share %": ":.2f"})
    charts.append(save_chart(fig, 2, "state_opportunity_share", "State Category Opportunity and Parle Value Share", chart_dir))

    # 3. Opportunity/right-to-win bubble matrix.
    cat_nd = max_metric(category, ["State"], nd).rename(columns={nd: "Category ND %"})
    pa_nd = sum_metric(parle, ["State"], nd).rename(columns={nd: "Parle ND %"})
    matrix = state.merge(cat_nd, on="State", how="left").merge(pa_nd, on="State", how="left").fillna(0)
    matrix["ND Gap pp"] = (matrix["Category ND %"] - matrix["Parle ND %"]).clip(lower=0)
    matrix["Bubble"] = matrix["Category Value"].clip(lower=0)
    fig = px.scatter(matrix, x="Category Value", y="Parle Share %", size="Bubble", color="ND Gap pp", text="State", color_continuous_scale="RdYlGn_r", hover_data={"Category ND %": ":.2f", "Parle ND %": ":.2f"})
    fig.update_traces(textposition="top center")
    charts.append(save_chart(fig, 3, "growth_priority_matrix", "Growth Prioritization: Opportunity × Current Right-to-Win", chart_dir))

    # 4. State-category share heatmap.
    cat_sc = sum_metric(category, ["State", "S4CATEGORY"], sales).rename(columns={sales: "Category Value"})
    pa_sc = sum_metric(parle, ["State", "S4CATEGORY"], sales).rename(columns={sales: "Parle Value"})
    heat = cat_sc.merge(pa_sc, on=["State", "S4CATEGORY"], how="left").fillna({"Parle Value": 0})
    heat["Share %"] = 100 * heat["Parle Value"] / heat["Category Value"]
    pivot = heat.pivot(index="State", columns="S4CATEGORY", values="Share %")
    fig = px.imshow(pivot, text_auto=".1f", aspect="auto", color_continuous_scale="YlOrRd", labels={"color": "Parle share %"})
    charts.append(save_chart(fig, 4, "state_category_share_heatmap", "Parle Value Share by State and Category", chart_dir))

    # 5. Revenue-state size and share.
    revenue_states = ["Gujarat", "Telangana", "Punjab", "Haryana"]
    rgm = state[state["State"].isin(revenue_states)].sort_values("Category Value", ascending=False)
    fig = make_subplots(specs=[[{"secondary_y": True}]])
    fig.add_bar(x=rgm["State"], y=rgm["Category Value"], name="Category value", marker_color="#7F4E2A", secondary_y=False)
    fig.add_scatter(x=rgm["State"], y=rgm["Parle Share %"], name="Parle share %", mode="lines+markers+text", text=rgm["Parle Share %"].round(2), textposition="top center", marker_color="#F5A623", secondary_y=True)
    fig.update_yaxes(title_text="Category value", secondary_y=False)
    fig.update_yaxes(title_text="Parle share %", secondary_y=True)
    charts.append(save_chart(fig, 5, "revenue_states_size_share", "Revenue-Led States: Category Size and Share Potential", chart_dir))

    # 6. Distribution gap and store headroom by state.
    cat_nd_sc = max_metric(category, ["State", "S4CATEGORY"], nd).rename(columns={nd: "Category ND %"})
    pa_nd_sc = sum_metric(parle, ["State", "S4CATEGORY"], nd).rename(columns={nd: "Parle ND %"})
    store_sc = max_metric(category, ["State", "S4CATEGORY"], stores).rename(columns={stores: "Category Stores"})
    gaps = cat_nd_sc.merge(pa_nd_sc, on=["State", "S4CATEGORY"], how="left").merge(store_sc, on=["State", "S4CATEGORY"], how="left").fillna(0)
    gaps["ND Gap pp"] = (gaps["Category ND %"] - gaps["Parle ND %"]).clip(lower=0)
    gaps["Store Headroom"] = gaps["ND Gap pp"] * gaps["Category Stores"] / 100
    state_gap = gaps.groupby("State", as_index=False).agg(**{"Average ND Gap pp": ("ND Gap pp", "mean"), "Store Headroom": ("Store Headroom", "sum")})
    fig = make_subplots(rows=1, cols=2, subplot_titles=("Average numeric-distribution gap", "Estimated store headroom"))
    fig.add_bar(x=state_gap["State"], y=state_gap["Average ND Gap pp"], marker_color="#2F5597", name="ND gap", row=1, col=1)
    fig.add_bar(x=state_gap["State"], y=state_gap["Store Headroom"], marker_color="#70AD47", name="Store headroom", row=1, col=2)
    fig.update_xaxes(tickangle=-45)
    charts.append(save_chart(fig, 6, "distribution_headroom", "Distribution Gap and Addressable Store Headroom", chart_dir))

    # 7. National pack-format mix.
    pack = sum_metric(sku, ["Pack Bucket"], sales).sort_values(sales, ascending=False)
    fig = px.pie(pack, names="Pack Bucket", values=sales, hole=0.45, color_discrete_sequence=px.colors.qualitative.Safe)
    fig.update_traces(textposition="inside", textinfo="percent+label")
    charts.append(save_chart(fig, 7, "pack_format_mix", "Parle Sales Mix by Pack Format", chart_dir))

    # 8. Pack mix by state.
    pack_state = sum_metric(sku, ["State", "Pack Bucket"], sales)
    pack_state["Mix %"] = 100 * pack_state[sales] / pack_state.groupby("State")[sales].transform("sum")
    fig = px.bar(pack_state, x="State", y="Mix %", color="Pack Bucket", barmode="stack", hover_data={sales: ":,.2f"}, color_discrete_sequence=px.colors.qualitative.Safe)
    fig.update_xaxes(tickangle=-35)
    charts.append(save_chart(fig, 8, "pack_mix_by_state", "Pack-Format Mix by State", chart_dir))

    # 9. Brand value contribution.
    brand = sum_metric(sku, ["Brand"], sales).nlargest(15, sales).sort_values(sales)
    fig = px.bar(brand, x=sales, y="Brand", orientation="h", color=sales, color_continuous_scale="Blues", text_auto=".3s")
    charts.append(save_chart(fig, 9, "brand_value", "Top Parle Brands by Sales Value", chart_dir))

    # 10. GTM-state active brand and SKU depth.
    active = sku[sku[sales].fillna(0).gt(0)]
    depth = active.groupby("State", as_index=False).agg(**{"Active Brands": ("Brand", "nunique"), "Active SKUs": ("ITEM", "nunique")})
    fig = make_subplots(rows=1, cols=2, subplot_titles=("Active brands", "Active SKUs"))
    fig.add_bar(x=depth["State"], y=depth["Active Brands"], marker_color="#4472C4", name="Brands", row=1, col=1)
    fig.add_bar(x=depth["State"], y=depth["Active SKUs"], marker_color="#ED7D31", name="SKUs", row=1, col=2)
    fig.update_xaxes(tickangle=-45)
    charts.append(save_chart(fig, 10, "portfolio_depth", "Portfolio Breadth and Range Depth by State", chart_dir))

    # 11. Price ladder using price-per-unit and SKU value.
    priced = sku.dropna(subset=["Price per Sales Unit", sales]).copy()
    priced["Price Band"] = pd.cut(priced["Price per Sales Unit"], [-math.inf, 10, 15, 20, 30, 50, math.inf], labels=["≤₹10", "₹10–15", "₹15–20", "₹20–30", "₹30–50", "₹50+"])
    ladder = sum_metric(priced, ["Price Band"], sales)
    fig = px.bar(ladder, x="Price Band", y=sales, color="Price Band", text_auto=".3s", color_discrete_sequence=px.colors.sequential.Oranges)
    charts.append(save_chart(fig, 11, "price_ladder", "Current Portfolio Value by Price Ladder", chart_dir))

    # 12. Distribution versus out-of-stock diagnostic.
    diag = parle.groupby("State", as_index=False).agg(**{"Numeric Distribution %": (nd, "sum"), "Numeric OOS %": ("Numeric Out of Stock (C)", "mean"), "Sales Value": (sales, "sum")}).dropna()
    fig = px.scatter(diag, x="Numeric Distribution %", y="Numeric OOS %", size="Sales Value", color="State", text="State", hover_data={"Sales Value": ":,.2f"})
    fig.update_traces(textposition="top center")
    charts.append(save_chart(fig, 12, "distribution_oos", "Distribution Reach vs. Out-of-Stock Diagnostic", chart_dir))

    links = "\n".join(f'<li><a href="{html.escape(item["file"])}">{html.escape(item["title"])}</a></li>' for item in charts)
    index = f'''<!doctype html><html><head><meta charset="utf-8"><title>PAPL Interactive Charts</title>
<style>body{{font-family:Arial;max-width:980px;margin:40px auto;color:#222}}h1{{color:#17365D}}li{{margin:14px 0}}a{{color:#2F5597;font-size:18px}}</style></head>
<body><h1>PAPL Commercial Intelligence – Interactive Chart Review</h1><p>Final Ver · Apr–Sep 2025 · {len(frame):,} populated rows</p><ol>{links}</ol></body></html>'''
    (chart_dir / "index.html").write_text(index, encoding="utf-8")
    return charts


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("input", nargs="?", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    parser.add_argument("--charts", action="store_true", help="Generate interactive Plotly charts")
    parser.add_argument("--charts-only", action="store_true", help="Skip JSON profiling and generate charts")
    args = parser.parse_args()

    if args.charts or args.charts_only:
        print(f"Loading populated rows from {args.sheet}...", flush=True)
        frame = load_final_data(args.input, args.sheet)
        print(f"Creating charts from {len(frame):,} populated rows...", flush=True)
        charts = create_interactive_charts(frame, CHART_DIR)
        print(f"Created {len(charts)} interactive charts in {CHART_DIR}", flush=True)
        if args.charts_only:
            return

    workbook = load_workbook(args.input, read_only=True, data_only=True)
    if args.sheet not in workbook.sheetnames:
        raise ValueError(f"Sheet {args.sheet!r} not found. Available sheets: {workbook.sheetnames}")
    result = {
        "file": str(args.input),
        "file_size_mb": round(args.input.stat().st_size / 1024 / 1024, 2),
        "sheets": {},
    }
    worksheet = workbook[args.sheet]
    print(f"Profiling {worksheet.title} ({worksheet.max_row - 1:,} data rows)...", flush=True)
    result["sheets"][worksheet.title] = profile_sheet(worksheet)
    workbook.close()

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(result, indent=2), encoding="utf-8")
    print(f"Profile written to {args.output}")


if __name__ == "__main__":
    main()
