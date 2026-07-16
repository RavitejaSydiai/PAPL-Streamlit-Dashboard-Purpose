"""Validate five SKU records directly against Original Ver."""
from pathlib import Path
from collections import defaultdict
import sys
import pandas as pd
from openpyxl import load_workbook
sys.path.insert(0,str(Path(__file__).resolve().parents[1]))
from src.data.sku_productivity_data import ND_SOURCE,ROS_SOURCE,STATE_NAMES

ROOT=Path(__file__).resolve().parents[1]; DATA=ROOT/"data"/"processed"/"sku_productivity.parquet"; EXCEL=ROOT/"data"/"PAPL Phase 1- Transposed File_3_vs.xlsx"
def main():
    d=pd.read_parquet(DATA); selected=d.sort_values("ND_ROS_Score",ascending=False).groupby("State").head(1).head(5); targets=set(zip(selected.State,selected.ITEM)); raw=defaultdict(dict)
    wb=load_workbook(EXCEL,read_only=True,data_only=True); ws=wb["Original Ver"]; rows=ws.iter_rows(values_only=True); h=list(next(rows)); idx={v:i for i,v in enumerate(h)}
    for r in rows:
        if r[idx["MANUFACTURER"]]!="PARLE AGRO" or not r[idx["ITEM"]]:continue
        state=STATE_NAMES.get(str(r[idx["Markets"]]).split("/")[-2]); key=(state,str(r[idx["ITEM"]]).strip())
        if key not in targets:continue
        if isinstance(r[idx[ND_SOURCE]],(int,float)):raw[key]["ND"]=float(r[idx[ND_SOURCE]])
        if isinstance(r[idx[ROS_SOURCE]],(int,float)):raw[key]["ROS"]=float(r[idx[ROS_SOURCE]])
    wb.close(); lines=["# Five-SKU Original Ver validation",""]; passed=0
    for _,r in selected.iterrows():
        v=raw[(r.State,r.ITEM)]; ok=abs(v["ND"]-r.Numeric_Distribution)<1e-9 and abs(v["ROS"]-r.Unweighted_ROS)<1e-9 and abs(r.ND_ROS_Score-v["ND"]*v["ROS"])<1e-9; passed+=ok
        lines += [f"## {r.State} · {r.ITEM}",f"- Excel ND: {v['ND']:.6f}; processed ND: {r.Numeric_Distribution:.6f}",f"- Excel ROS: {v['ROS']:.6f}; processed ROS: {r.Unweighted_ROS:.6f}",f"- Score check: {r.ND_ROS_Score:.6f} = ND × ROS",f"- Result: {'PASS' if ok else 'FAIL'}",""]
    lines += [f"Validation passed: {passed}/5"] ; (ROOT/"docs"/"sku_productivity_validation.md").write_text("\n".join(lines),encoding="utf-8"); print(f"SKU validation: {passed}/5 passed")
    if passed!=5:raise SystemExit(1)
if __name__=="__main__":main()
