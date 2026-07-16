"""Original Ver-only loader and tidy fact aggregator."""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path
import sys

import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
from config.fact_aggregation_rules import FACT_RULES

SHEET_NAME = "Original Ver"
PERIOD_COLUMN = "Apr 25 - Sep 25"
STATE_NAMES = {"AP":"Andhra Pradesh","DEL":"Delhi","GUJ":"Gujarat","HAR":"Haryana","KAR":"Karnataka","MAH":"Maharashtra","PUNJ":"Punjab","RAJ":"Rajasthan","TLG":"Telangana","TN":"Tamil Nadu","UP":"Uttar Pradesh","WB":"West Bengal"}


def _state(market):
    if not market:
        return None
    code = str(market).split("/")[-2] if "/" in str(market) else str(market)
    return STATE_NAMES.get(code, code)


def load_original_comparison(path: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Return named-manufacturer facts and blank-manufacturer benchmarks."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError("Original Ver sheet not found")
    ws = workbook[SHEET_NAME]
    rows = ws.iter_rows(values_only=True)
    headers = [str(v).strip() if v is not None else "" for v in next(rows)]
    idx = {name: i for i, name in enumerate(headers)}
    required = {"Markets", "S4CATEGORY", "MANUFACTURER", PERIOD_COLUMN, *FACT_RULES}
    missing = required - set(headers)
    if missing:
        raise ValueError(f"Required Original Ver columns missing: {sorted(missing)}")
    accum = defaultdict(lambda: [0.0, 0, None])
    for row in rows:
        category = row[idx["S4CATEGORY"]]
        if category is None:
            continue
        manufacturer = row[idx["MANUFACTURER"]]
        manufacturer = str(manufacturer).strip() if manufacturer is not None else "__CATEGORY_TOTAL__"
        state = _state(row[idx["Markets"]])
        for fact, rule in FACT_RULES.items():
            value = row[idx[fact]]
            if not isinstance(value, (int, float)):
                continue
            key = (state, str(category).strip(), manufacturer, fact)
            cell = accum[key]
            cell[0] += float(value); cell[1] += 1
            cell[2] = float(value) if cell[2] is None else max(cell[2], float(value))
    workbook.close()
    records = []
    for (state, category, manufacturer, fact), (total, count, maximum) in accum.items():
        rule = FACT_RULES[fact]
        value = total if rule["aggregation"] == "sum" else maximum if rule["aggregation"] == "max" else total / count
        records.append({"State": state, "S4CATEGORY": category, "Manufacturer": manufacturer, "Fact": fact, "Metric_Value": value, "Observation_Count": count, "Aggregation": rule["aggregation"], "Fact_Class": rule["class"]})
    tidy = pd.DataFrame(records)
    blank = tidy[tidy["Manufacturer"].eq("__CATEGORY_TOTAL__")].copy()
    named = tidy[~tidy["Manufacturer"].eq("__CATEGORY_TOTAL__")].copy()
    return named, blank
