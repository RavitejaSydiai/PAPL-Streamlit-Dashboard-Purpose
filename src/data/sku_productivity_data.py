"""Original Ver loader for Parle Agro State × ITEM ND/ROS facts."""
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

SHEET_NAME="Original Ver"; PERIOD_COLUMN="Apr 25 - Sep 25"; PARLE="PARLE AGRO"
ND_SOURCE="Numeric Distribution Handling (C)"; ROS_SOURCE="Unwghtd ROS (VOL UNIT CASES) Handling (C)"
STATE_NAMES={"AP":"Andhra Pradesh","DEL":"Delhi","GUJ":"Gujarat","HAR":"Haryana","KAR":"Karnataka","MAH":"Maharashtra","PUNJ":"Punjab","RAJ":"Rajasthan","TLG":"Telangana","TN":"Tamil Nadu","UP":"Uttar Pradesh","WB":"West Bengal"}

def _state(market): return STATE_NAMES.get(str(market).split("/")[-2])
def build_sku_productivity_source(path:Path)->pd.DataFrame:
    wb=load_workbook(path,read_only=True,data_only=True); ws=wb[SHEET_NAME]; rows=ws.iter_rows(values_only=True); headers=list(next(rows)); idx={v:i for i,v in enumerate(headers)}
    required=["Markets","ITEM","MANUFACTURER",ND_SOURCE,ROS_SOURCE,PERIOD_COLUMN]
    missing=[c for c in required if c not in idx]
    if missing: raise ValueError(f"Missing exact Original Ver columns: {missing}")
    records=[]
    for row in rows:
        if row[idx["MANUFACTURER"]]!=PARLE or not row[idx["ITEM"]]: continue
        state=_state(row[idx["Markets"]]); item=str(row[idx["ITEM"]]).strip()
        for fact in [ND_SOURCE,ROS_SOURCE]:
            value=row[idx[fact]]
            if isinstance(value,(int,float)): records.append({"State":state,"ITEM":item,"Fact":fact,"Metric_Value":float(value)})
    wb.close(); long=pd.DataFrame(records)
    duplicates=long.duplicated(["State","ITEM","Fact"],keep=False)
    if duplicates.any(): raise ValueError(f"Unexpected duplicate State-ITEM-Fact observations: {duplicates.sum()}")
    wide=long.pivot(index=["State","ITEM"],columns="Fact",values="Metric_Value").reset_index()
    return wide.rename(columns={ND_SOURCE:"Numeric_Distribution",ROS_SOURCE:"Unweighted_ROS"})

def load_processed(path:Path)->pd.DataFrame: return pd.read_parquet(path)
