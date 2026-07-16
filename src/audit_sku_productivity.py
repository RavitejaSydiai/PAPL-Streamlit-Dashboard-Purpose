"""Audit the Original Ver grain for Parle ND–ROS SKU productivity analysis."""
from collections import Counter,defaultdict
from pathlib import Path
import json
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[1]; INPUT=ROOT/"data"/"PAPL Phase 1- Transposed File_3_vs.xlsx"; OUTPUT=ROOT/"docs"/"sku_productivity_data_audit.json"
ND="Numeric Distribution Handling (C)"; ROS="Unwghtd ROS (VOL UNIT CASES) Handling (C)"
STATES={"AP":"Andhra Pradesh","DEL":"Delhi","GUJ":"Gujarat","HAR":"Haryana","KAR":"Karnataka","MAH":"Maharashtra","PUNJ":"Punjab","RAJ":"Rajasthan","TLG":"Telangana","TN":"Tamil Nadu","UP":"Uttar Pradesh","WB":"West Bengal"}

def main():
    wb=load_workbook(INPUT,read_only=True,data_only=True); ws=wb["Original Ver"]; rows=ws.iter_rows(values_only=True); h=list(next(rows)); i={v:n for n,v in enumerate(h)}
    values=defaultdict(list); dimension_patterns=defaultdict(set); source_rows=0
    for r in rows:
        if r[i["MANUFACTURER"]]!="PARLE AGRO" or not r[i["ITEM"]]: continue
        state=STATES.get(str(r[i["Markets"]]).split("/")[-2]); item=str(r[i["ITEM"]]).strip()
        for fact in [ND,ROS]:
            value=r[i[fact]]
            if isinstance(value,(int,float)):
                source_rows+=1; values[(state,item,fact)].append(float(value)); dimension_patterns[(state,item,fact)].add((r[i["ITEM - Pack Size"]],r[i["Attribute"]],r[i["Addition"]]))
    wb.close(); counts=Counter(len(v) for v in values.values()); duplicate={k:v for k,v in values.items() if len(v)>1}; conflicting={k:v for k,v in duplicate.items() if len(set(v))>1}
    nd_values=[x for (s,it,f),v in values.items() if f==ND for x in v]; ros_values=[x for (s,it,f),v in values.items() if f==ROS for x in v]
    paired=len(set((s,it) for s,it,f in values if f==ND)&set((s,it) for s,it,f in values if f==ROS))
    result={"sheet":"Original Ver","manufacturer":"PARLE AGRO","period":"Apr 25 - Sep 25","source_metric_rows":source_rows,"state_item_fact_keys":len(values),"observations_per_key":dict(counts),"duplicate_keys":len(duplicate),"conflicting_duplicate_keys":len(conflicting),"paired_state_items":paired,"nd_min":min(nd_values),"nd_max":max(nd_values),"ros_min":min(ros_values),"ros_max":max(ros_values),"duplicate_dimension_variants":sum(len(dimension_patterns[k])>1 for k in duplicate),"examples":[{"State":k[0],"ITEM":k[1],"Fact":k[2],"Values":v,"DimensionVariants":len(dimension_patterns[k])} for k,v in list(conflicting.items())[:20]]}
    OUTPUT.write_text(json.dumps(result,indent=2),encoding="utf-8"); print(json.dumps(result,indent=2))
if __name__=="__main__": main()
