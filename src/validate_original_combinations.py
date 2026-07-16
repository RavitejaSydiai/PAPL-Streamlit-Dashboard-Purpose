"""Independent Excel-to-Parquet validation for five State × S4CATEGORY combinations."""
from collections import defaultdict
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[1]
CHECKS=[("Andhra Pradesh","MANGO DRINKS"),("Delhi","MANGO DRINKS"),("Gujarat","PACKAGED PLAIN WATER"),("Maharashtra","MILK DRINKS"),("Punjab","RTD SUGAR")]
CODES={"AP":"Andhra Pradesh","DEL":"Delhi","GUJ":"Gujarat","MAH":"Maharashtra","PUNJ":"Punjab"}

def main():
    wb=load_workbook(ROOT/"data"/"PAPL Phase 1- Transposed File_3_vs.xlsx",read_only=True,data_only=True); ws=wb["Original Ver"]
    rows=ws.iter_rows(values_only=True); h=list(next(rows)); idx={v:i for i,v in enumerate(h)}; raw=defaultdict(float)
    for r in rows:
        m=r[idx["MANUFACTURER"]]; c=r[idx["S4CATEGORY"]]; v=r[idx["Sales Value (000)"]]; market=r[idx["Markets"]]
        if m is None or c is None or not isinstance(v,(int,float)): continue
        code=str(market).split("/")[-2]; state=CODES.get(code)
        if (state,c) in CHECKS: raw[(state,c,str(m).strip())]+=float(v)
    wb.close(); wide=pd.read_parquet(ROOT/"data"/"processed"/"manufacturer_comparison_wide.parquet"); lines=["# Manual validation: five Original Ver combinations",""]
    passed=0
    for state,cat in CHECKS:
        g=wide[(wide.State==state)&(wide.S4CATEGORY==cat)].dropna(subset=["Sales Value (000)"]); p=g[g.Manufacturer.eq("PARLE AGRO")]
        processed=float(p.iloc[0]["Sales Value (000)"]) if not p.empty else None; source=raw.get((state,cat,"PARLE AGRO")); ok=processed is not None and abs(processed-source)<1e-6; passed+=ok
        leader=max(((v,m) for (s,c,m),v in raw.items() if s==state and c==cat),default=(None,None))
        lines += [f"## {state} · {cat}",f"- Raw Excel Parle Sales Value sum: {source:,.6f}",f"- Processed Parquet Parle Sales Value: {processed:,.6f}",f"- Raw Excel leader: {leader[1]} ({leader[0]:,.6f})",f"- Result: {'PASS' if ok else 'FAIL'}",""]
    lines += [f"Validated combinations passed: {passed}/{len(CHECKS)}"]
    (ROOT/"docs"/"manual_validation.md").write_text("\n".join(lines),encoding="utf-8"); print(f"Manual validation: {passed}/{len(CHECKS)} passed")
    if passed != len(CHECKS): raise SystemExit(1)
if __name__=="__main__": main()
