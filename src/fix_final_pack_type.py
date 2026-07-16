"""Correct Final Ver Pack Type values from the authoritative Pack Type_Size field."""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path("data/PAPL Phase 1- Transposed File_3_vs.xlsx")
TEMP_OUTPUT = WORKBOOK.with_name(WORKBOOK.stem + "_correcting.xlsx")


def derive_pack_type(value):
    if value is None or str(value).strip() == "":
        return None
    match = re.match(r"\s*([A-Za-z]+)", str(value))
    return match.group(1).upper() if match else None


def main():
    workbook = load_workbook(WORKBOOK)
    worksheet = workbook["Final Ver"]
    headers = {cell.value: cell.column for cell in worksheet[1]}
    pack_type_column = headers["Pack Type"]
    pack_size_column = headers["Pack Type_Size"]
    changed = 0
    populated = 0

    for row_number in range(2, worksheet.max_row + 1):
        pack_size = worksheet.cell(row_number, pack_size_column).value
        if pack_size is None:
            continue
        populated += 1
        expected = derive_pack_type(pack_size)
        target = worksheet.cell(row_number, pack_type_column)
        if target.value != expected:
            target.value = expected
            changed += 1

    workbook.save(TEMP_OUTPUT)
    workbook.close()
    print(f"Pack Type_Size rows checked: {populated:,}")
    print(f"Pack Type values corrected: {changed:,}")
    print(f"Corrected workbook written to: {TEMP_OUTPUT}")


if __name__ == "__main__":
    main()
