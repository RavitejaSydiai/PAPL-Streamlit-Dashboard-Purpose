"""Audit the transformation from Original Ver to Final Ver."""

from __future__ import annotations

import json
import math
import re
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


INPUT = Path("data/PAPL Phase 1- Transposed File_3_vs.xlsx")
OUTPUT = Path("docs/original_vs_final_audit.json")
METRICS = [
    "Sales Value (000)", "Sales (VOL UNIT CASES)", "Sales Units",
    "Share of Sales Value - Product",
    "Share of Sales - Sales (VOL UNIT CASES) - Product",
    "Share of Sales - Sales Units - Product",
    "Stocks Forward (VOL UNIT CASES)", "Stocks Backward (VOL UNIT CASES)",
    "Price per Sales (VOL UNIT CASES)",
    "Sales (VOL UNIT CASES) Price Index - Product", "Price per Sales Unit",
    "Stocks (VOL UNIT CASES)", "Stocks (VOL UNIT CASES) Shr - Product",
    "Purchase (VOL UNIT CASES) Shr - Product",
    "Numeric Distribution Handling (C)", "Numeric Out of Stock (C)",
    "Wghtd Dist Out of Stock", "Number of Stores Retailing (C)",
    "Relative Numeric Distribution Handling (C) - Product",
    "Unwghtd ROS (VOL UNIT CASES) Handling (C)",
    "Value Shr in Handlers (C) - Product - CATEGORY",
    "Wghtd Dist Handling (C) - CATEGORY", "Sales (Vol (KG/LT/000NO))",
    "Purchase (VOL UNIT CASES)", "Days of Supply (Vol (KG/LT/000NO))",
]


def populated_records(ws, predicate=None):
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))
    index = {name: position for position, name in enumerate(headers)}
    records = []
    for row in rows:
        if not any(value is not None and value != "" for value in row):
            continue
        if predicate is None or predicate(row, index):
            records.append(dict(zip(headers, row)))
    return records


def same(left, right):
    if left is None and right is None:
        return True
    if isinstance(left, (int, float)) and isinstance(right, (int, float)):
        return math.isclose(float(left), float(right), rel_tol=1e-12, abs_tol=1e-9)
    return left == right


def expected_size(pack):
    if not pack:
        return None
    match = re.search(r"(\d+(?:\.\d+)?)\s*ML", str(pack), flags=re.I)
    return float(match.group(1)) if match else None


def expected_pack_type(pack):
    if not pack:
        return None
    match = re.match(r"\s*([A-Z]+)", str(pack))
    return match.group(1) if match else None


def main():
    workbook = load_workbook(INPUT, read_only=True, data_only=True)
    original = populated_records(
        workbook["Original Ver"],
        lambda row, idx: row[idx["S4CATEGORY"]] is not None
        and (row[idx["MANUFACTURER"]] is None or row[idx["MANUFACTURER"]] == "PARLE AGRO"),
    )
    final = populated_records(workbook["Final Ver"])
    workbook.close()

    common = ["Markets", "S4CATEGORY", "MANUFACTURER", "ITEM", "Attribute", *METRICS]
    def signature(row):
        return tuple(row.get(column) for column in common)

    original_signatures = Counter(signature(row) for row in original)
    final_signatures = Counter(signature(row) for row in final)
    missing_from_final = sum((original_signatures - final_signatures).values())
    added_in_final = sum((final_signatures - original_signatures).values())

    metric_summary = {}
    for metric in METRICS:
        old_values = [row[metric] for row in original if isinstance(row.get(metric), (int, float))]
        new_values = [row[metric] for row in final if isinstance(row.get(metric), (int, float))]
        metric_summary[metric] = {
            "original_count": len(old_values), "final_count": len(new_values),
            "original_sum": round(sum(old_values), 6), "final_sum": round(sum(new_values), 6),
        }

    item_pack_counts = {}
    for row in original:
        item = row.get("ITEM")
        pack = row.get("ITEM - Pack Size")
        if item:
            item_pack_counts.setdefault(item, Counter())[pack] += 1
    item_pack = {item: counts.most_common(1)[0][0] for item, counts in item_pack_counts.items()}

    pack_size_mismatch = 0
    size_mismatch = 0
    pack_type_mismatch = 0
    pack_type_pairs = Counter()
    auditable_pack_rows = 0
    for right in final:
        item = right.get("ITEM")
        original_pack = item_pack.get(item) if item else None
        if item:
            auditable_pack_rows += 1
        if item and not same(original_pack, right.get("Pack Type_Size")):
            pack_size_mismatch += 1
        expected = expected_size(original_pack)
        actual = right.get("Size") if isinstance(right.get("Size"), (int, float)) else None
        if item and not same(expected, actual):
            size_mismatch += 1
        expected_type = expected_pack_type(original_pack)
        actual_type = right.get("Pack Type")
        if item and not same(expected_type, actual_type):
            pack_type_mismatch += 1
            pack_type_pairs[(str(expected_type), str(actual_type))] += 1

    result = {
        "filtered_original_rows": len(original),
        "final_populated_rows": len(final),
        "row_count_equal": len(original) == len(final),
        "common_row_multiset": {
            "missing_from_final": missing_from_final,
            "added_in_final": added_in_final,
            "exactly_equal": missing_from_final == 0 and added_in_final == 0,
        },
        "metric_summary": metric_summary,
        "pack_field_audit": {
            "auditable_item_rows": auditable_pack_rows,
            "pack_type_size_mismatches": pack_size_mismatch,
            "size_mismatches": size_mismatch,
            "pack_type_mismatches": pack_type_mismatch,
            "top_pack_type_mismatch_pairs": [
                {"expected": expected, "actual": actual, "count": count}
                for (expected, actual), count in pack_type_pairs.most_common(20)
            ],
        },
    }
    OUTPUT.write_text(json.dumps(result, indent=2), encoding="utf-8")
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
