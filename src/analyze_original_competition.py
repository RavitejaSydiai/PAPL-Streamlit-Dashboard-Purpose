"""Competitive analysis using the complete hidden Original Ver sheet."""

from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


INPUT = Path("data/PAPL Phase 1- Transposed File_3_vs.xlsx")
OUTPUT = Path("docs/original_competitive_analysis.json")
STATE_NAMES = {
    "AP": "Andhra Pradesh", "DEL": "Delhi", "GUJ": "Gujarat", "HAR": "Haryana",
    "KAR": "Karnataka", "MAH": "Maharashtra", "PUNJ": "Punjab", "RAJ": "Rajasthan",
    "TLG": "Telangana", "TN": "Tamil Nadu", "UP": "Uttar Pradesh", "WB": "West Bengal",
}
SALES = "Sales Value (000)"


def main():
    workbook = load_workbook(INPUT, read_only=True, data_only=True)
    ws = workbook["Original Ver"]
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))
    idx = {name: i for i, name in enumerate(headers)}
    keep = ["Markets", "S4CATEGORY", "MANUFACTURER", "ITEM", SALES]
    records = []
    for row in rows:
        value = row[idx[SALES]]
        category = row[idx["S4CATEGORY"]]
        if category is None or not isinstance(value, (int, float)):
            continue
        records.append({column: row[idx[column]] for column in keep})
    workbook.close()
    df = pd.DataFrame(records)
    df["State"] = df["Markets"].str.extract(r"/([^/]+)/INDIA$")[0].map(STATE_NAMES)
    df["Brand"] = df["ITEM"].fillna("").str.lstrip("#").str.split("\\", regex=False).str[0].str.strip()
    df.loc[df["Brand"].eq(""), "Brand"] = pd.NA
    df["Hierarchy"] = "Category total"
    df.loc[df["MANUFACTURER"].notna(), "Hierarchy"] = "Manufacturer aggregate"
    df.loc[df["ITEM"].notna(), "Hierarchy"] = "SKU"

    category_rows = df[df["Hierarchy"].eq("Category total")]
    manufacturer_rows = df[df["Hierarchy"].eq("Manufacturer aggregate")]
    sku_rows = df[df["Hierarchy"].eq("SKU")]

    national_category = category_rows.groupby("S4CATEGORY", as_index=False)[SALES].sum().rename(columns={SALES: "Category Value"})
    company_category = manufacturer_rows.groupby(["S4CATEGORY", "MANUFACTURER"], as_index=False)[SALES].sum().rename(columns={SALES: "Manufacturer Value"})
    company_category = company_category.merge(national_category, on="S4CATEGORY", how="left")
    company_category["Value Share %"] = 100 * company_category["Manufacturer Value"] / company_category["Category Value"]
    company_category["Rank"] = company_category.groupby("S4CATEGORY")["Manufacturer Value"].rank(method="min", ascending=False).astype(int)

    state_category = category_rows.groupby(["State", "S4CATEGORY"], as_index=False)[SALES].sum().rename(columns={SALES: "Category Value"})
    state_company = manufacturer_rows.groupby(["State", "S4CATEGORY", "MANUFACTURER"], as_index=False)[SALES].sum().rename(columns={SALES: "Manufacturer Value"})
    state_company = state_company.merge(state_category, on=["State", "S4CATEGORY"], how="left")
    state_company["Value Share %"] = 100 * state_company["Manufacturer Value"] / state_company["Category Value"]
    state_company["Rank"] = state_company.groupby(["State", "S4CATEGORY"])["Manufacturer Value"].rank(method="min", ascending=False).astype(int)

    national_company = manufacturer_rows.groupby("MANUFACTURER", as_index=False)[SALES].sum().sort_values(SALES, ascending=False)
    national_company["Rank"] = range(1, len(national_company) + 1)
    total_market = national_category["Category Value"].sum()
    national_company["NARTD Value Share %"] = 100 * national_company[SALES] / total_market

    brand = sku_rows.groupby("Brand", as_index=False)[SALES].sum().sort_values(SALES, ascending=False)
    brand["Rank"] = range(1, len(brand) + 1)

    parle_categories = company_category[company_category["MANUFACTURER"].eq("PARLE AGRO")].sort_values("Manufacturer Value", ascending=False)
    parle_states = state_company[state_company["MANUFACTURER"].eq("PARLE AGRO")].sort_values(["State", "Manufacturer Value"], ascending=[True, False])
    parle_national = national_company[national_company["MANUFACTURER"].eq("PARLE AGRO")]

    top_by_category = {}
    for category, group in company_category.groupby("S4CATEGORY"):
        top_by_category[category] = group.nsmallest(10, "Rank")[["MANUFACTURER", "Manufacturer Value", "Value Share %", "Rank"]].to_dict("records")

    result = {
        "sales_value_rows_analyzed": len(df),
        "total_category_value": float(total_market),
        "parle_national": parle_national.to_dict("records"),
        "parle_by_category": parle_categories.to_dict("records"),
        "parle_by_state_category": parle_states.to_dict("records"),
        "top_manufacturers_national": national_company.head(20).to_dict("records"),
        "top_brands_national": brand.head(20).to_dict("records"),
        "top_manufacturers_by_category": top_by_category,
    }
    OUTPUT.write_text(json.dumps(result, indent=2), encoding="utf-8")
    company_category.to_csv("docs/original_company_category.csv", index=False)
    state_company.to_csv("docs/original_state_company_category.csv", index=False)
    brand.to_csv("docs/original_brand_ranking.csv", index=False)
    print(f"Analyzed {len(df):,} Sales Value rows; output written to {OUTPUT}")


if __name__ == "__main__":
    main()
